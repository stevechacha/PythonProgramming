import openpyxl as xl

wb = xl.load_workbook('Transaction.xlsx')
print(wb)
sheet = wb['Sheet1']
print(sheet)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell)
